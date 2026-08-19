import re
import chromadb
from sentence_transformers import SentenceTransformer
from openpyxl import load_workbook


EXCEL_FILE = "UTI_evaluation_dataset_v2.xlsx"

TOP_K_RETRIEVAL = 10
TOP_K_FINAL = 5

LEXICAL_WEIGHT = 0.10
TITLE_WEIGHT = 0.05


# ============================================================
# DATASET
# ============================================================

wb = load_workbook(
    EXCEL_FILE,
    read_only=True,
    data_only=True
)

ws = wb["Evaluation_Set"]

test_cases = []

for row in ws.iter_rows(min_row=2, values_only=True):

    question = row[1]
    relevant_sources = row[2]

    if question and relevant_sources:

        test_cases.append(
            (
                question,
                set(str(relevant_sources).split("|"))
            )
        )

print("Questions loaded:", len(test_cases))


# ============================================================
# MPNet
# ============================================================

print()
print("Loading MPNet...")

model = SentenceTransformer(
    "all-mpnet-base-v2"
)

print("MPNet loaded")


# ============================================================
# DATABASE
# ============================================================

client = chromadb.PersistentClient(
    path="./chroma_db_mpnet"
)

collection = client.get_collection(
    name="uti_guideline_mpnet"
)

print("Documents:", collection.count())


# ============================================================
# TOKENIZATION
# ============================================================

STOPWORDS = {
    "the", "a", "an", "and", "or", "of", "to", "for", "with",
    "what", "which", "how", "should", "be", "is", "are", "was",
    "were", "do", "does", "did", "when", "where", "who", "that",
    "this", "these", "those", "in", "on", "at", "from", "by",
    "about", "all", "people", "person", "patients", "patient"
}


def tokenize(text):

    return {
        word
        for word in re.findall(
            r"[a-zA-Z0-9]+",
            text.lower()
        )
        if len(word) >= 3
        and word not in STOPWORDS
    }


def lexical_overlap(query, document):

    q = tokenize(query)
    d = tokenize(document)

    if not q or not d:
        return 0.0

    overlap = len(q.intersection(d))

    precision = overlap / len(q)
    recall = overlap / len(d)

    if precision + recall == 0:
        return 0.0

    return (
        2 * precision * recall
        / (precision + recall)
    )


# ============================================================
# CONSERVATIVE TITLE SIGNAL
# ============================================================

def title_signal(query, metadata):

    title = str(
        metadata.get("title", "")
    ).lower()

    q = query.lower()

    score = 0.0

    # Exact recommendation/table identity cues.
    # No generic word-by-word title bonus.
    strong_phrases = [
        "when prescribing antibiotic",
        "prescribing antibiotic treatment",
        "microbiological results",
        "non-pregnant women aged 16 years and over",
        "pregnant women aged 12 years and over",
        "men aged 16 years and over",
        "children and young people under 16 years",
        "pregnant women and men",
    ]

    for phrase in strong_phrases:

        if phrase in q and phrase in title:
            score += 1.0

    # Specific age/group cues.
    group_phrases = [
        "non-pregnant",
        "pregnant",
        "men",
        "children",
        "under 16",
        "16 years and over",
    ]

    for phrase in group_phrases:

        if phrase in q and phrase in title:
            score += 0.25

    return min(score, 1.5)


# ============================================================
# METRICS
# ============================================================

def calculate_metrics(ranks):

    total = len(ranks)

    hits = [
        r for r in ranks
        if r is not None
    ]

    recall5 = len(hits) / total

    hit1 = sum(
        1 for r in ranks
        if r == 1
    ) / total

    mrr = sum(
        1 / r
        for r in hits
    ) / total

    return recall5, hit1, mrr


# ============================================================
# PRECOMPUTE
# ============================================================

print()
print("Precomputing retrieval...")

prepared = []

for index, (question, relevant_sources) in enumerate(
    test_cases,
    start=1
):

    embedding = model.encode(
        [question],
        normalize_embeddings=True
    )[0].tolist()

    results = collection.query(
        query_embeddings=[embedding],
        n_results=TOP_K_RETRIEVAL
    )

    candidates = []

    for document, metadata, distance in zip(
        results["documents"][0],
        results["metadatas"][0],
        results["distances"][0]
    ):

        similarity = 1 - distance

        lexical = lexical_overlap(
            question,
            document
        )

        title = title_signal(
            question,
            metadata
        )

        candidates.append({
            "metadata": metadata,
            "similarity": similarity,
            "lexical": lexical,
            "title": title
        })

    prepared.append(
        (
            question,
            relevant_sources,
            candidates
        )
    )

    print(
        f"\rPrepared {index}/{len(test_cases)}",
        end=""
    )

print()


# ============================================================
# BASELINE HYBRID VS TITLE-AWARE HYBRID
# ============================================================

baseline_ranks = []
title_aware_ranks = []

for question, relevant_sources, candidates in prepared:

    # Current proven hybrid:
    # similarity + 0.10 * lexical overlap

    baseline = sorted(
        candidates,
        key=lambda x:
            x["similarity"]
            + LEXICAL_WEIGHT * x["lexical"],
        reverse=True
    )

    baseline_rank = None

    for rank, item in enumerate(
        baseline[:TOP_K_FINAL],
        start=1
    ):

        if item["metadata"]["source_id"] in relevant_sources:
            baseline_rank = rank
            break

    baseline_ranks.append(
        baseline_rank
    )

    # New title-aware hybrid:
    # add only a small title signal.

    title_aware = sorted(
        candidates,
        key=lambda x:
            x["similarity"]
            + LEXICAL_WEIGHT * x["lexical"]
            + TITLE_WEIGHT * x["title"],
        reverse=True
    )

    title_rank = None

    for rank, item in enumerate(
        title_aware[:TOP_K_FINAL],
        start=1
    ):

        if item["metadata"]["source_id"] in relevant_sources:
            title_rank = rank
            break

    title_aware_ranks.append(
        title_rank
    )


# ============================================================
# RESULTS
# ============================================================

b = calculate_metrics(
    baseline_ranks
)

t = calculate_metrics(
    title_aware_ranks
)

print()
print()
print("=" * 70)
print("CURRENT HYBRID VS TITLE-AWARE HYBRID")
print("=" * 70)

print()
print(
    "                         CURRENT        TITLE-AWARE"
)
print("-" * 70)

print(
    f"Recall@5               {b[0]:.4f}"
    f"          {t[0]:.4f}"
)

print(
    f"Hit@1                  {b[1]:.4f}"
    f"          {t[1]:.4f}"
)

print(
    f"MRR                    {b[2]:.4f}"
    f"          {t[2]:.4f}"
)

print()
print("-" * 70)

print(
    "Recall@5 change:",
    f"{t[0] - b[0]:+.4f}"
)

print(
    "Hit@1 change:",
    f"{t[1] - b[1]:+.4f}"
)

print(
    "MRR change:",
    f"{t[2] - b[2]:+.4f}"
)

print()
print("Lexical weight:", LEXICAL_WEIGHT)
print("Title weight:", TITLE_WEIGHT)

print()
print("Evaluation completed.")
