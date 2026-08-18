import re
import chromadb
from sentence_transformers import SentenceTransformer
from openpyxl import load_workbook

# ============================================================
# REGRESSION TEST — FINAL HYBRID V3
# ============================================================

EXCEL_FILE = "UTI_evaluation_dataset_v2.xlsx"
MODEL_NAME = "all-mpnet-base-v2"
DATABASE_PATH = "./chroma_db_mpnet"
COLLECTION_NAME = "uti_guideline_mpnet"

TOP_K_RETRIEVAL = 10
TOP_K_FINAL = 5
LEXICAL_WEIGHT = 0.10
SIMILARITY_THRESHOLD = 0.20

# Locked benchmark from the current best retrieval version
BASELINE_RECALL = 1.0000
BASELINE_HIT1 = 0.9636
BASELINE_MRR = 0.9773

STOPWORDS = {
    "the","a","an","and","or","of","to","for","with","what","which",
    "how","should","be","is","are","was","were","do","does","did",
    "when","where","who","that","this","these","those","in","on",
    "at","from","by","about","all","people","person","patients","patient"
}

def tokenize(text):
    return {
        w for w in re.findall(r"[a-zA-Z0-9]+", text.lower())
        if len(w) >= 3 and w not in STOPWORDS
    }

def lexical_overlap(q, d):
    q_tokens = tokenize(q)
    d_tokens = tokenize(d)

    if not q_tokens or not d_tokens:
        return 0.0

    overlap = len(q_tokens & d_tokens)
    precision = overlap / len(q_tokens)
    recall = overlap / len(d_tokens)

    if precision + recall == 0:
        return 0.0

    return 2 * precision * recall / (precision + recall)

def metrics(ranks):
    n = len(ranks)

    recall5 = sum(
        r is not None and r <= 5
        for r in ranks
    ) / n

    hit1 = sum(
        r == 1
        for r in ranks
    ) / n

    mrr = sum(
        1 / r if r is not None else 0
        for r in ranks
    ) / n

    return recall5, hit1, mrr


print("Loading evaluation dataset...")

wb = load_workbook(
    EXCEL_FILE,
    read_only=True,
    data_only=True
)

ws = wb["Evaluation_Set"]

test_cases = []

for row in ws.iter_rows(min_row=2, values_only=True):
    question = row[1]
    relevant = row[2]

    if question and relevant:
        test_cases.append(
            (question, set(str(relevant).split("|")))
        )

print("Questions loaded:", len(test_cases))

print()
print("Loading MPNet...")

model = SentenceTransformer(MODEL_NAME)

client = chromadb.PersistentClient(
    path=DATABASE_PATH
)

collection = client.get_collection(
    name=COLLECTION_NAME
)

print("Documents:", collection.count())

ranks = []
failed_cases = []

print()
print("=" * 70)
print("RUNNING 55-QUESTION REGRESSION TEST")
print("=" * 70)

for index, (question, relevant_sources) in enumerate(
    test_cases,
    start=1
):

    embedding = model.encode(
        [question],
        normalize_embeddings=True
    )[0].tolist()

    results = collection.query(
        query_embeddings=[embedding],
        n_results=collection.count()
    )

    candidates = []

    for document, metadata, distance in zip(
        results["documents"][0],
        results["metadatas"][0],
        results["distances"][0]
    ):

        similarity = 1 - distance

        # IMPORTANT:
        # The threshold is applied to semantic similarity,
        # while ranking still uses the locked Hybrid score.
        if similarity < SIMILARITY_THRESHOLD:
            continue

        lexical = lexical_overlap(
            question,
            document
        )

        hybrid = (
            similarity
            + LEXICAL_WEIGHT * lexical
        )

        candidates.append({
            "source_id": metadata["source_id"],
            "similarity": similarity,
            "hybrid": hybrid
        })

    candidates.sort(
        key=lambda x: x["hybrid"],
        reverse=True
    )

    top5 = candidates[:TOP_K_FINAL]

    rank = None

    for position, item in enumerate(top5, start=1):
        if item["source_id"] in relevant_sources:
            rank = position
            break

    ranks.append(rank)

    if rank is None:
        failed_cases.append({
            "index": index,
            "question": question,
            "expected": sorted(relevant_sources),
            "top5": [
                x["source_id"]
                for x in top5
            ]
        })

    print(
        f"Prepared {index}/{len(test_cases)}",
        end="\r"
    )

recall5, hit1, mrr = metrics(ranks)

print()
print()
print("=" * 70)
print("REGRESSION RESULTS")
print("=" * 70)

print()
print(f"Threshold: {SIMILARITY_THRESHOLD:.2f}")
print(f"Lexical weight: {LEXICAL_WEIGHT:.2f}")
print()

print(f"Recall@5 : {recall5:.4f}   (locked: {BASELINE_RECALL:.4f})")
print(f"Hit@1    : {hit1:.4f}   (locked: {BASELINE_HIT1:.4f})")
print(f"MRR      : {mrr:.4f}   (locked: {BASELINE_MRR:.4f})")

print()
print("=" * 70)
print("REGRESSION CHECK")
print("=" * 70)

TOLERANCE = 0.0001

checks = {
    "Recall@5": recall5 >= BASELINE_RECALL - TOLERANCE,
    "Hit@1": hit1 >= BASELINE_HIT1 - TOLERANCE,
    "MRR": mrr >= BASELINE_MRR - TOLERANCE
}

all_pass = True

for name, passed in checks.items():
    print(
        f"{name:<10}:",
        "PASS" if passed else "FAIL"
    )

    if not passed:
        all_pass = False

print()

if failed_cases:
    print(
        "Failed retrieval cases:",
        len(failed_cases)
    )

    for item in failed_cases:
        print()
        print(
            f"[{item['index']}/{len(test_cases)}]"
        )
        print("Question:", item["question"])
        print(
            "Expected:",
            " | ".join(item["expected"])
        )
        print(
            "Top-5:",
            " | ".join(item["top5"])
        )
else:
    print("Failed retrieval cases: 0")

print()
print("=" * 70)

if all_pass and not failed_cases:
    print("FINAL REGRESSION: PASS")
else:
    print("FINAL REGRESSION: FAIL")

print("=" * 70)
