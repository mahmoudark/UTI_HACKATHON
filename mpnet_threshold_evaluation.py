import re
import chromadb
from sentence_transformers import SentenceTransformer
from openpyxl import load_workbook


# ============================================================
# SETTINGS
# ============================================================

EXCEL_FILE = "UTI_evaluation_dataset_v2.xlsx"

MODEL_NAME = "all-mpnet-base-v2"

DATABASE_PATH = "./chroma_db_mpnet"
COLLECTION_NAME = "uti_guideline_mpnet"

TOP_K_RETRIEVAL = 10
TOP_K_FINAL = 5

LEXICAL_WEIGHT = 0.10


# ============================================================
# LOAD DATASET
# ============================================================

print("Loading evaluation dataset...")

wb = load_workbook(
    EXCEL_FILE,
    read_only=True,
    data_only=True
)

ws = wb["Evaluation_Set"]

test_cases = []

for row in ws.iter_rows(min_row=2, values_only=True):

    question = row[1]
    relevant_sources = row[2]

    if question and relevant_sources:

        sources = set(
            str(relevant_sources).split("|")
        )

        test_cases.append(
            (question, sources)
        )

print(
    "Questions loaded:",
    len(test_cases)
)


# ============================================================
# LOAD MPNet
# ============================================================

print()
print("Loading MPNet...")

model = SentenceTransformer(
    MODEL_NAME
)

print("MPNet loaded")


# ============================================================
# LOAD DATABASE
# ============================================================

client = chromadb.PersistentClient(
    path=DATABASE_PATH
)

collection = client.get_collection(
    name=COLLECTION_NAME
)

print("Database loaded")
print("Documents:", collection.count())


# ============================================================
# TOKENIZATION
# ============================================================

STOPWORDS = {
    "the", "a", "an", "and", "or", "of", "to", "for", "with",
    "what", "which", "how", "should", "be", "is", "are", "was",
    "were", "do", "does", "did", "when", "where", "who", "that",
    "this", "these", "those", "in", "on", "at", "from", "by",
    "about", "all", "people", "person", "patients", "patient"
}


def tokenize(text):

    words = re.findall(
        r"[a-zA-Z0-9]+",
        text.lower()
    )

    return {
        word
        for word in words
        if len(word) >= 3
        and word not in STOPWORDS
    }


def lexical_overlap(query, document):

    query_tokens = tokenize(query)
    document_tokens = tokenize(document)

    if not query_tokens or not document_tokens:
        return 0.0

    overlap = len(
        query_tokens.intersection(document_tokens)
    )

    precision = overlap / len(query_tokens)
    recall = overlap / len(document_tokens)

    if precision + recall == 0:
        return 0.0

    return (
        2 * precision * recall
        / (precision + recall)
    )


# ============================================================
# PRECOMPUTE RETRIEVAL
# ============================================================

print()
print("Precomputing MPNet + Hybrid retrieval...")

prepared = []

for index, (question, relevant_sources) in enumerate(
    test_cases,
    start=1
):

    embedding = model.encode(
        [question],
        normalize_embeddings=True
    )[0].tolist()

    results = collection.query(
        query_embeddings=[embedding],
        n_results=collection.count()
    )

    candidates = []

    for document, metadata, distance in zip(
        results["documents"][0],
        results["metadatas"][0],
        results["distances"][0]
    ):

        similarity = 1 - distance

        lexical_score = lexical_overlap(
            question,
            document
        )

        hybrid_score = (
            similarity
            + LEXICAL_WEIGHT * lexical_score
        )

        candidates.append({
            "source_id": metadata["source_id"],
            "similarity": similarity,
            "lexical_score": lexical_score,
            "hybrid_score": hybrid_score
        })

    candidates.sort(
        key=lambda x: x["hybrid_score"],
        reverse=True
    )

    prepared.append(
        (
            question,
            relevant_sources,
            candidates
        )
    )

    print(
        f"Prepared {index}/{len(test_cases)}"
    )


# ============================================================
# THRESHOLD EVALUATION
# ============================================================

def evaluate_threshold(threshold):

    answered = 0
    correct = 0

    ranks = []

    for question, relevant_sources, candidates in prepared:

        selected = [
            item
            for item in candidates
            if item["similarity"] >= threshold
        ]

        selected = selected[:TOP_K_FINAL]

        if not selected:
            ranks.append(None)
            continue

        answered += 1

        found_rank = None

        for rank, item in enumerate(
            selected,
            start=1
        ):

            if item["source_id"] in relevant_sources:

                found_rank = rank
                break

        ranks.append(found_rank)

        if found_rank is not None:
            correct += 1


    total = len(prepared)

    recall = correct / total

    coverage = answered / total

    hit1 = sum(
        1
        for rank in ranks
        if rank == 1
    ) / total

    mrr = sum(
        1 / rank
        for rank in ranks
        if rank is not None
    ) / total

    return (
        recall,
        coverage,
        hit1,
        mrr,
        ranks
    )


# ============================================================
# RUN TESTS
# ============================================================

print()
print("=" * 75)
print("MPNET + HYBRID THRESHOLD EVALUATION")
print("=" * 75)

print()
print(
    "Dataset:",
    len(test_cases),
    "questions"
)

print(
    "Model:",
    MODEL_NAME
)

print(
    "Lexical weight:",
    LEXICAL_WEIGHT
)

print()
print(
    "Threshold      Recall@5     Coverage     Hit@1       MRR"
)

print("-" * 75)


thresholds = [
    0.20,
    0.22,
    0.24,
    0.25,
    0.26,
    0.28,
    0.30,
    0.32,
    0.35,
    0.40,
    0.45,
    0.50
]


results = []

for threshold in thresholds:

    recall, coverage, hit1, mrr, ranks = evaluate_threshold(
        threshold
    )

    results.append(
        (
            threshold,
            recall,
            coverage,
            hit1,
            mrr
        )
    )

    print(
        f"{threshold:<14.2f}"
        f"{recall:<13.4f}"
        f"{coverage:<13.4f}"
        f"{hit1:<12.4f}"
        f"{mrr:.4f}"
    )


# ============================================================
# SAFE THRESHOLDS
# ============================================================

safe = [
    result
    for result in results
    if result[1] >= 1.0
]


print()
print("=" * 75)
print("THRESHOLDS WITH 100% RECALL@5")
print("=" * 75)

if safe:

    for threshold, recall, coverage, hit1, mrr in safe:

        print(
            f"Threshold {threshold:.2f} | "
            f"Coverage {coverage:.4f} | "
            f"Hit@1 {hit1:.4f} | "
            f"MRR {mrr:.4f}"
        )

    best_safe = max(
        safe,
        key=lambda x: (
            x[1],
            x[3],
            x[4],
            x[2]
        )
    )

    print()
    print(
        "Recommended threshold candidate:",
        best_safe[0]
    )

else:

    print(
        "No tested threshold preserved 100% Recall@5."
    )

    print(
        "Do NOT increase the production threshold yet."
    )


print()
print("=" * 75)
print("EVALUATION COMPLETED")
print("=" * 75)