import re
import chromadb
from sentence_transformers import SentenceTransformer
from openpyxl import load_workbook

EXCEL_FILE = "UTI_evaluation_dataset_v2.xlsx"
TOP_K_RETRIEVAL = 10
TOP_K_FINAL = 5
LEXICAL_WEIGHT = 0.10

wb = load_workbook(EXCEL_FILE, read_only=True, data_only=True)
ws = wb["Evaluation_Set"]

test_cases = []
for row in ws.iter_rows(min_row=2, values_only=True):
    question, sources = row[1], row[2]
    if question and sources:
        test_cases.append((question, set(str(sources).split("|"))))

print("Questions loaded:", len(test_cases))
print("Loading MPNet...")
model = SentenceTransformer("all-mpnet-base-v2")
client = chromadb.PersistentClient(path="./chroma_db_mpnet")
collection = client.get_collection(name="uti_guideline_mpnet")
print("Documents:", collection.count())

STOPWORDS = {
    "the","a","an","and","or","of","to","for","with","what","which",
    "how","should","be","is","are","was","were","do","does","did",
    "when","where","who","that","this","these","those","in","on",
    "at","from","by","about","all","people","person","patients","patient"
}

def tokens(text):
    return {
        w for w in re.findall(r"[a-zA-Z0-9]+", text.lower())
        if len(w) >= 3 and w not in STOPWORDS
    }

def lexical_overlap(q, d):
    q, d = tokens(q), tokens(d)
    if not q or not d:
        return 0.0
    overlap = len(q & d)
    p, r = overlap / len(q), overlap / len(d)
    return 0.0 if p + r == 0 else 2 * p * r / (p + r)

def rank_metrics(ranks):
    n = len(ranks)
    recall5 = sum(r is not None and r <= 5 for r in ranks) / n
    hit1 = sum(r == 1 for r in ranks) / n
    mrr = sum(1/r if r is not None else 0 for r in ranks) / n
    return recall5, hit1, mrr

def precision5(top5, relevant):
    return sum(
        x["metadata"]["source_id"] in relevant for x in top5
    ) / len(top5)

baseline_ranks, hybrid_ranks = [], []
baseline_p, hybrid_p = [], []
changes = []

for i, (question, relevant) in enumerate(test_cases, 1):
    emb = model.encode([question], normalize_embeddings=True)[0].tolist()
    r = collection.query(query_embeddings=[emb], n_results=TOP_K_RETRIEVAL)

    candidates = []
    for doc, meta, dist in zip(
        r["documents"][0], r["metadatas"][0], r["distances"][0]
    ):
        sim = 1 - dist
        lex = lexical_overlap(question, doc)
        candidates.append({
            "metadata": meta,
            "similarity": sim,
            "lexical": lex,
            "hybrid": sim + LEXICAL_WEIGHT * lex
        })

    base5 = candidates[:TOP_K_FINAL]
    base_rank = next(
        (rank for rank, x in enumerate(base5, 1)
         if x["metadata"]["source_id"] in relevant),
        None
    )

    candidates.sort(key=lambda x: x["hybrid"], reverse=True)
    hyb5 = candidates[:TOP_K_FINAL]
    hyb_rank = next(
        (rank for rank, x in enumerate(hyb5, 1)
         if x["metadata"]["source_id"] in relevant),
        None
    )

    baseline_ranks.append(base_rank)
    hybrid_ranks.append(hyb_rank)
    baseline_p.append(precision5(base5, relevant))
    hybrid_p.append(precision5(hyb5, relevant))

    if base_rank != hyb_rank:
        changes.append((i, question, relevant, base_rank, hyb_rank,
                        [x["metadata"]["source_id"] for x in base5],
                        [x["metadata"]["source_id"] for x in hyb5]))

b = rank_metrics(baseline_ranks)
h = rank_metrics(hybrid_ranks)
bp = sum(baseline_p) / len(baseline_p)
hp = sum(hybrid_p) / len(hybrid_p)

print()
print("=" * 70)
print("FINAL MASTER EVALUATION")
print("=" * 70)
print("                         BASELINE       HYBRID")
print("-" * 70)
print(f"Precision@5            {bp:.4f}          {hp:.4f}")
print(f"Recall@5               {b[0]:.4f}          {h[0]:.4f}")
print(f"Hit@1                  {b[1]:.4f}          {h[1]:.4f}")
print(f"MRR                    {b[2]:.4f}          {h[2]:.4f}")
print("-" * 70)
print(f"Precision@5 change: {hp-bp:+.4f}")
print(f"Recall@5 change:    {h[0]-b[0]:+.4f}")
print(f"Hit@1 change:       {h[1]-b[1]:+.4f}")
print(f"MRR change:         {h[2]-b[2]:+.4f}")

print()
print("=" * 70)
print("NON-IDENTICAL CASES")
print("=" * 70)
for item in changes:
    i, q, expected, br, hr, bt, ht = item
    print()
    print(f"[{i}/{len(test_cases)}]")
    print("Question:", q)
    print("Expected:", " | ".join(sorted(expected)))
    print("Baseline rank:", br)
    print("Hybrid rank:", hr)
    print("Baseline Top-5:", " | ".join(bt))
    print("Hybrid Top-5:", " | ".join(ht))

print()
print("Total changed cases:", len(changes))
print("Evaluation completed.")
