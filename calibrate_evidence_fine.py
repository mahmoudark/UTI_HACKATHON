import chromadb
from sentence_transformers import SentenceTransformer
from openpyxl import load_workbook


EXCEL_FILE = "UTI_evaluation_dataset_v2.xlsx"
MODEL_NAME = "all-mpnet-base-v2"

client = chromadb.PersistentClient(
    path="./chroma_db_mpnet"
)

collection = client.get_collection(
    name="uti_guideline_mpnet"
)

model = SentenceTransformer(MODEL_NAME)

wb = load_workbook(
    EXCEL_FILE,
    read_only=True,
    data_only=True
)

ws = wb["Evaluation_Set"]


rows = []

for row in ws.iter_rows(min_row=2, values_only=True):

    question = row[1]
    relevant_sources = row[2]

    if not question or not relevant_sources:
        continue

    relevant = {
        x.strip()
        for x in str(relevant_sources).split("|")
    }

    embedding = model.encode(
        [question],
        normalize_embeddings=True
    )[0].tolist()

    result = collection.query(
        query_embeddings=[embedding],
        n_results=4
    )

    for rank, (meta, distance) in enumerate(
        zip(
            result["metadatas"][0],
            result["distances"][0]
        ),
        start=1
    ):

        cosine = 1 - (distance / 2)

        rows.append({
            "question": question,
            "rank": rank,
            "source_id": meta["source_id"],
            "cosine": cosine,
            "relevant": meta["source_id"] in relevant
        })


print()
print("=" * 90)
print("FINE COSINE CALIBRATION: 0.75 - 0.85")
print("=" * 90)

selected = [
    x
    for x in rows
    if 0.75 <= x["cosine"] <= 0.85
]

selected.sort(
    key=lambda x: x["cosine"],
    reverse=True
)

print()
print(
    f"{'Cosine':<10}"
    f"{'Rank':<7}"
    f"{'Relevant':<11}"
    f"{'Source':<12}"
    f"Question"
)

print("-" * 90)

for x in selected:
    print(
        f"{x['cosine']:<10.4f}"
        f"{x['rank']:<7}"
        f"{str(x['relevant']):<11}"
        f"{x['source_id']:<12}"
        f"{x['question']}"
    )


print()
print("=" * 90)
print("CUMULATIVE RELEVANCE")
print("=" * 90)

for threshold in [
    0.70,
    0.71,
    0.72,
    0.73,
    0.74,
    0.75,
    0.76,
    0.77,
    0.78,
    0.79,
    0.80,
    0.81,
    0.82,
    0.83,
    0.84,
    0.85
]:

    above = [
        x
        for x in rows
        if x["cosine"] >= threshold
    ]

    if not above:
        continue

    relevant = sum(
        x["relevant"]
        for x in above
    )

    rate = relevant / len(above)

    print(
        f"Cosine >= {threshold:.2f} | "
        f"Samples={len(above):2d} | "
        f"Relevant={relevant:2d} | "
        f"Match={rate * 100:6.2f}%"
    )