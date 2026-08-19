import chromadb
from sentence_transformers import SentenceTransformer, CrossEncoder
from openpyxl import load_workbook


# ============================================================
# SETTINGS
# ============================================================

EXCEL_FILE = "UTI_evaluation_dataset.xlsx"

TOP_K_RETRIEVAL = 10
TOP_K_FINAL = 5


# ============================================================
# LOAD DATASET
# ============================================================

print("Loading evaluation dataset...")

wb = load_workbook(
    EXCEL_FILE,
    read_only=True,
    data_only=True
)

ws = wb["Evaluation_Set"]

test_cases = []

for row in ws.iter_rows(min_row=2, values_only=True):

    question = row[1]
    expected_source = row[2]

    if question and expected_source:

        relevant_sources = set(
            str(expected_source).split("|")
        )

        test_cases.append(
            (question, relevant_sources)
        )
print(
    "Questions loaded:",
    len(test_cases)
)


# ============================================================
# LOAD MPNet
# ============================================================

print()
print("Loading MPNet...")

embedding_model = SentenceTransformer(
    "all-mpnet-base-v2"
)

print("MPNet loaded")


# ============================================================
# LOAD RERANKER
# ============================================================

print()
print("Loading Cross-Encoder...")

reranker = CrossEncoder(
    "cross-encoder/ms-marco-MiniLM-L-6-v2"
)

print("Reranker loaded")


# ============================================================
# LOAD CHROMA
# ============================================================

client = chromadb.PersistentClient(
    path="./chroma_db_mpnet"
)

collection = client.get_collection(
    name="uti_guideline_mpnet"
)

print(
    "Documents:",
    collection.count()
)


# ============================================================
# METRIC FUNCTION
# ============================================================

def calculate_metrics(ranks):

    total = len(ranks)

    hits = [
        r for r in ranks
        if r is not None
    ]

    recall_at_5 = (
        len(hits) / total
    )

    hit_at_1 = (
        sum(
            1 for r in ranks
            if r == 1
        ) / total
    )

    mrr = (
        sum(
            1 / r
            for r in hits
        ) / total
    )

    return (
        recall_at_5,
        hit_at_1,
        mrr
    )


# ============================================================
# EVALUATE
# ============================================================

baseline_ranks = []
reranker_ranks = []


print()
print("=" * 70)
print("EVALUATING 55 QUESTIONS")
print("=" * 70)


for index, (question, expected) in enumerate(
    test_cases,
    start=1
):

    print()
    print(
        f"[{index}/{len(test_cases)}]"
    )

    print(
        "Question:",
        question
    )

    print(
        "Expected:",
        expected
    )


    # ========================================================
    # MPNet RETRIEVAL
    # ========================================================

    query_embedding = embedding_model.encode(
        [question],
        normalize_embeddings=True
    )[0].tolist()


    results = collection.query(
        query_embeddings=[query_embedding],
        n_results=TOP_K_RETRIEVAL
    )


    documents = results["documents"][0]
    metadatas = results["metadatas"][0]


    # ========================================================
    # BASELINE RANK
    # ========================================================

    baseline_rank = None

    for rank, metadata in enumerate(
        metadatas[:TOP_K_FINAL],
        start=1
    ):

        if metadata["source_id"] in expected:

            baseline_rank = rank

            break

    baseline_ranks.append(
        baseline_rank
    )

    # ========================================================
    # RERANK
    # ========================================================

    pairs = [
        (question, document)
        for document in documents
    ]


    scores = reranker.predict(
        pairs
    )


    candidates = []


    for document, metadata, score in zip(
        documents,
        metadatas,
        scores
    ):

        candidates.append(
            (
                float(score),
                metadata
            )
        )


    candidates.sort(
        key=lambda x: x[0],
        reverse=True
    )


    reranker_rank = None


    for rank, (_, metadata) in enumerate(
        candidates[:TOP_K_FINAL],
        start=1
    ):

        if metadata["source_id"] in expected:

            reranker_rank = rank

            break


    reranker_ranks.append(
        reranker_rank
    )


    print(
        "Baseline rank:",
        baseline_rank
    )

    print(
        "Reranker rank:",
        reranker_rank
    )


# ============================================================
# FINAL METRICS
# ============================================================

baseline_metrics = calculate_metrics(
    baseline_ranks
)

reranker_metrics = calculate_metrics(
    reranker_ranks
)


# ============================================================
# RESULTS
# ============================================================

print()
print()
print("=" * 70)
print("FINAL 55-QUESTION EVALUATION")
print("=" * 70)


print()
print("                 BASELINE       RERANKER")
print("-" * 70)


print(
    f"Recall@5       {baseline_metrics[0]:.4f}"
    f"          {reranker_metrics[0]:.4f}"
)


print(
    f"Hit@1          {baseline_metrics[1]:.4f}"
    f"          {reranker_metrics[1]:.4f}"
)


print(
    f"MRR            {baseline_metrics[2]:.4f}"
    f"          {reranker_metrics[2]:.4f}"
)


print()
print("=" * 70)


# ============================================================
# IMPROVEMENT
# ============================================================

mrr_improvement = (
    reranker_metrics[2]
    -
    baseline_metrics[2]
)


hit1_improvement = (
    reranker_metrics[1]
    -
    baseline_metrics[1]
)


print(
    "MRR improvement:",
    f"{mrr_improvement:+.4f}"
)


print(
    "Hit@1 improvement:",
    f"{hit1_improvement:+.4f}"
)


print()
print("Evaluation completed.")