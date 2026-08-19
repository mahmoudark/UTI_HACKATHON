import re
import chromadb
from sentence_transformers import SentenceTransformer
from openpyxl import load_workbook

EXCEL_FILE = "UTI_evaluation_dataset_v2.xlsx"
TOP_K = 5

wb = load_workbook(EXCEL_FILE, read_only=True, data_only=True)
ws = wb["Evaluation_Set"]

test_cases = []

for row in ws.iter_rows(min_row=2, values_only=True):
    question, sources = row[1], row[2]

    if question and sources:
        relevant = set(str(sources).split("|"))
        test_cases.append((question, relevant))

print("Questions loaded:", len(test_cases))

model = SentenceTransformer("all-mpnet-base-v2")

client = chromadb.PersistentClient(
    path="./chroma_db_mpnet"
)

collection = client.get_collection(
    name="uti_guideline_mpnet"
)

STOPWORDS = {
    "the","a","an","and","or","of","to","for","with","what","which",
    "how","should","be","is","are","was","were","do","does","did",
    "when","where","who","that","this","these","those","in","on",
    "at","from","by","about","all","people","person","patients","patient"
}


def tokens(text):
    return {
        w
        for w in re.findall(r"[a-zA-Z0-9]+", text.lower())
        if len(w) >= 3 and w not in STOPWORDS
    }


def lexical_overlap(q, d):
    q_tokens = tokens(q)
    d_tokens = tokens(d)

    if not q_tokens or not d_tokens:
        return 0.0

    overlap = len(q_tokens & d_tokens)

    precision = overlap / len(q_tokens)
    recall = overlap / len(d_tokens)

    if precision + recall == 0:
        return 0.0

    return 2 * precision * recall / (precision + recall)


print()
print("=" * 75)
print("PRECISION@5 DIAGNOSTIC")
print("=" * 75)

distribution = {}

for i, (question, relevant) in enumerate(test_cases, 1):

    embedding = model.encode(
        [question],
        normalize_embeddings=True
    )[0].tolist()

    results = collection.query(
        query_embeddings=[embedding],
        n_results=10
    )

    candidates = []

    for doc, metadata, distance in zip(
        results["documents"][0],
        results["metadatas"][0],
        results["distances"][0]
    ):

        similarity = 1 - distance
        lexical = lexical_overlap(question, doc)

        hybrid = similarity + 0.10 * lexical

        candidates.append({
            "source_id": metadata["source_id"],
            "similarity": similarity,
            "hybrid": hybrid
        })

    candidates.sort(
        key=lambda x: x["hybrid"],
        reverse=True
    )

    top5 = candidates[:5]

    relevant_count = sum(
        item["source_id"] in relevant
        for item in top5
    )

    precision = relevant_count / 5

    distribution[precision] = distribution.get(precision, 0) + 1

    print()
    print(f"[{i}/{len(test_cases)}]")
    print("Question:", question)
    print("Expected:", " | ".join(sorted(relevant)))
    print(
        "Top-5:",
        " | ".join(item["source_id"] for item in top5)
    )
    print(
        f"Relevant in Top-5: {relevant_count}/5"
    )
    print(
        f"Precision@5: {precision:.2f}"
    )


print()
print("=" * 75)
print("PRECISION@5 DISTRIBUTION")
print("=" * 75)

for score in sorted(distribution):
    print(
        f"{score:.2f} -> {distribution[score]} questions"
    )