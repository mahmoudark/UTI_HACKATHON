import re
from statistics import mean

from openpyxl import load_workbook

from answer_engine_v3 import answer_with_guard


EXCEL_FILE = "UTI_evaluation_dataset_v2.xlsx"


# ------------------------------------------------------------
# Load evaluation set
# ------------------------------------------------------------

wb = load_workbook(
    EXCEL_FILE,
    read_only=True,
    data_only=True
)

ws = wb["Evaluation_Set"]

test_cases = []

for row in ws.iter_rows(min_row=2, values_only=True):

    question = row[1]
    sources = row[2]

    if question and sources:

        relevant = {
            x.strip()
            for x in str(sources).split("|")
        }

        test_cases.append(
            (question, relevant)
        )


# ------------------------------------------------------------
# Claim splitting
# ------------------------------------------------------------

def split_claims(answer):

    if not answer:
        return []

    text = re.sub(
        r"\s+",
        " ",
        answer.strip()
    )

    # Split on bullets, sentence endings, or semicolons.
    parts = re.split(
        r"(?:\n+|(?<=[.!?])\s+|;\s+|•\s*)",
        text
    )

    claims = []

    for part in parts:

        part = part.strip()

        if len(part) < 15:
            continue

        # Remove common template wrappers.
        if part.startswith("According to recommendation"):
            continue

        claims.append(part)

    return claims


# ------------------------------------------------------------
# Token normalization
# ------------------------------------------------------------

STOPWORDS = {
    "the", "a", "an", "and", "or", "of", "to", "for",
    "with", "what", "which", "how", "should", "be",
    "is", "are", "was", "were", "do", "does", "did",
    "when", "where", "who", "that", "this", "these",
    "those", "in", "on", "at", "from", "by", "about",
    "all", "people", "person", "patients", "patient",
    "according", "recommendation", "table", "guideline"
}


def tokens(text):

    return {
        w
        for w in re.findall(
            r"[a-zA-Z0-9]+",
            text.lower()
        )
        if len(w) >= 4
        and w not in STOPWORDS
    }


# ------------------------------------------------------------
# Claim support
# ------------------------------------------------------------

def claim_support(claim, evidence):

    claim_tokens = tokens(claim)
    evidence_tokens = tokens(evidence)

    if not claim_tokens:
        return 1.0

    if not evidence_tokens:
        return 0.0

    overlap = len(
        claim_tokens & evidence_tokens
    )

    return overlap / len(claim_tokens)


# ------------------------------------------------------------
# Evaluation
# ------------------------------------------------------------

print("=" * 75)
print("DAY 4 CLAIM-LEVEL FAITHFULNESS DIAGNOSTIC")
print("=" * 75)

print("Questions:", len(test_cases))
print()

all_claim_scores = []
answered_count = 0
refused_count = 0


for i, (question, relevant) in enumerate(
    test_cases,
    start=1
):

    result = answer_with_guard(question)

    if result["status"] == "REFUSED":

        refused_count += 1

        print(
            f"[{i:02}/{len(test_cases)}] REFUSED"
        )

        continue

    answered_count += 1

    answer = result["answer"]
    evidence = ""

    # Use the selected source document as the grounding text.
    if result.get("results"):

        best = result["results"][0]

        evidence = str(
            best.get("document", "")
        )

    claims = split_claims(answer)

    print()
    print(
        f"[{i:02}/{len(test_cases)}]"
    )

    print(
        "Source:",
        result["source"].get("source_id")
        if result.get("source")
        else "N/A"
    )

    print(
        "Claims:",
        len(claims)
    )

    question_scores = []

    for n, claim in enumerate(
        claims,
        start=1
    ):

        score = claim_support(
            claim,
            evidence
        )

        question_scores.append(score)
        all_claim_scores.append(score)

        label = (
            "SUPPORTED"
            if score >= 0.75
            else "UNSUPPORTED"
        )

        print(
            f"  Claim {n}: {label} "
            f"({score:.2f})"
        )

        print(
            "   ",
            claim
        )

    if question_scores:

        print(
            "Question Faithfulness Proxy:",
            f"{mean(question_scores):.4f}"
        )


# ------------------------------------------------------------
# Final numbers
# ------------------------------------------------------------

faithfulness_proxy = (
    mean(all_claim_scores)
    if all_claim_scores
    else 0.0
)


print()
print("=" * 75)
print("FINAL RESULTS")
print("=" * 75)

print(
    f"Answered:                  {answered_count}"
)

print(
    f"Refused:                   {refused_count}"
)

print(
    f"Claims evaluated:          {len(all_claim_scores)}"
)

print(
    f"Faithfulness Proxy:        {faithfulness_proxy:.4f}"
)

print()

print(
    "NOTE: This is a claim-level lexical "
    "faithfulness proxy, not human-annotated "
    "gold-standard Faithfulness."
)

print(
    "Day 4 target: 0.9000+"
)