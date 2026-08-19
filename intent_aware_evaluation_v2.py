PS D:\UTI_HACKATHON> Get-Content intent_aware_evaluation.py
import re
import chromadb
from sentence_transformers import SentenceTransformer
from openpyxl import load_workbook


# ============================================================
# SETTINGS
# ============================================================

EXCEL_FILE = "UTI_evaluation_dataset_v2.xlsx"

MODEL_NAME = "all-mpnet-base-v2"

DATABASE_PATH = "./chroma_db_mpnet"

COLLECTION_NAME = "uti_guideline_mpnet"

TOP_K_RETRIEVAL = 10

TOP_K_FINAL = 5

LEXICAL_WEIGHT = 0.10

# Small intent bonus.
# This bonus is only activated when the query strongly
# indicates a specific clinical intent.
INTENT_WEIGHT = 0.10


# ============================================================
# LOAD DATASET
# ============================================================

print("Loading evaluation dataset...")

wb = load_workbook(
    EXCEL_FILE,
    read_only=True,
    data_only=True
)

ws = wb["Evaluation_Set"]

test_cases = []

for row in ws.iter_rows(
    min_row=2,
    values_only=True
):

    question = row[1]

    expected = row[2]

    if question and expected:

        relevant_sources = set(
            str(expected).split("|")
        )

        test_cases.append(
            (
                question,
                relevant_sources
            )
        )

print(
    "Questions loaded:",
    len(test_cases)
)


# ============================================================
# LOAD MPNet + DATABASE
# ============================================================

print()

print("Loading MPNet...")

model = SentenceTransformer(
    MODEL_NAME
)

print("MPNet loaded")


client = chromadb.PersistentClient(
    path=DATABASE_PATH
)

collection = client.get_collection(
    name=COLLECTION_NAME
)

print(
    "Documents:",
    collection.count()
)


# ============================================================
# TEXT NORMALIZATION
# ============================================================

STOPWORDS = {
    "the",
    "a",
    "an",
    "and",
    "or",
    "of",
    "to",
    "for",
    "with",
    "what",
    "which",
    "how",
    "should",
    "be",
    "is",
    "are",
    "was",
    "were",
    "do",
    "does",
    "did",
    "when",
    "where",
    "who",
    "that",
    "this",
    "these",
    "those",
    "in",
    "on",
    "at",
    "from",
    "by",
    "about",
    "all",
    "people",
    "person",
    "patients",
    "patient"
}


def tokenize(text):

    return {
        word
        for word in re.findall(
            r"[a-zA-Z0-9]+",
            text.lower()
        )
        if len(word) >= 3
        and word not in STOPWORDS
    }


def lexical_overlap(
    query,
    document
):

    query_tokens = tokenize(
        query
    )

    document_tokens = tokenize(
        document
    )

    if (
        not query_tokens
        or not document_tokens
    ):
        return 0.0

    overlap = len(
        query_tokens.intersection(
            document_tokens
        )
    )

    precision = (
        overlap
        / len(query_tokens)
    )

    recall = (
        overlap
        / len(document_tokens)
    )

    if precision + recall == 0:

        return 0.0

    return (
        2 * precision * recall
        / (precision + recall)
    )


# ============================================================
# INTENT DETECTION
# ============================================================

def detect_intent(query):

    q = query.lower()

    # ========================================================
    # ANTIBIOTIC SELECTION / PRESCRIBING
    # ========================================================

    antibiotic_terms = [
        "antibiotic",
        "antibiotics",
        "antimicrobial",
        "antimicrobials"
    ]

    selection_terms = [
        "prescrib",
        "choos",
        "select",
        "choice",
        "choosing",
        "selecting",
        "selection",
        "decision",
        "decisions",
        "treatment choice"
    ]

    consideration_terms = [
        "consider",
        "considered",
        "consideration",
        "considerations",
        "factor",
        "factors",
        "guide",
        "guiding",
        "influence",
        "influences",
        "take into account",
        "taken into account",
        "account",
        "important",
        "before prescribing"
    ]

    has_antibiotic = any(
        term in q
        for term in antibiotic_terms
    )

    has_selection = any(
        term in q
        for term in selection_terms
    )

    has_consideration = any(
        term in q
        for term in consideration_terms
    )

    # Strong antibiotic-selection intent.
    if (
        has_antibiotic
        and has_selection
        and has_consideration
    ):
        return "ANTIBIOTIC_SELECTION"

    # Antibiotic prescribing/selection decisions
    # for lower UTI, even without explicit
    # consideration/factor wording.
    if (
        has_antibiotic
        and has_selection
        and (
            "lower uti" in q
            or
            "lower urinary tract infection" in q
        )
    ):
        return "ANTIBIOTIC_SELECTION"

    # Antibiotic + consideration/factor intent
    # for lower UTI.
    if (
        has_antibiotic
        and has_consideration
        and (
            "lower uti" in q
            or
            "lower urinary tract infection" in q
        )
    ):
        return "ANTIBIOTIC_SELECTION"

    # ========================================================
    # PATIENT GROUP / TABLE INTENTS
    # ========================================================

    if (
        "men aged 16 years and over" in q
        and "antibiotics" in q
    ):
        return "TABLE_3"

    if (
        "pregnant women aged 12 years and over"
        in q
        and "antibiotics" in q
    ):
        return "TABLE_2"

    if (
        "non-pregnant women aged 16 years and over"
        in q
        and "antibiotics" in q
    ):
        return "TABLE_1"

    if (
        "children and young people under 16 years"
        in q
        and "antibiotics" in q
    ):
        return "TABLE_4"

    return "GENERAL"


# ============================================================
# INTENT BONUS
# ============================================================

def intent_bonus(
    query,
    metadata,
    document
):

    intent = detect_intent(
        query
    )

    source_id = str(
        metadata.get(
            "source_id",
            ""
        )
    )

    document_lower = document.lower()


    # ========================================================
    # ANTIBIOTIC SELECTION
    # ========================================================

    if intent == "ANTIBIOTIC_SELECTION":

        # Exact target recommendation.
        if source_id == "1.4.1":

            return INTENT_WEIGHT


        # Weaker supporting signals.
        if (
            "when prescribing antibiotic treatment"
            in document_lower
        ):

            return (
                INTENT_WEIGHT
                * 0.50
            )


        if (
            "local antimicrobial resistance data"
            in document_lower
        ):

            return (
                INTENT_WEIGHT
                * 0.50
            )


    # ========================================================
    # TABLE 1
    # ========================================================

    if intent == "TABLE_1":

        if source_id == "TABLE_1":

            return INTENT_WEIGHT


    # ========================================================
    # TABLE 2
    # ========================================================

    if intent == "TABLE_2":

        if source_id == "TABLE_2":

            return INTENT_WEIGHT


    # ========================================================
    # TABLE 3
    # ========================================================

    if intent == "TABLE_3":

        if source_id == "TABLE_3":

            return INTENT_WEIGHT


    # ========================================================
    # TABLE 4
    # ========================================================

    if intent == "TABLE_4":

        if source_id == "TABLE_4":

            return INTENT_WEIGHT


    return 0.0


# ============================================================
# METRICS
# ============================================================

def calculate_metrics(
    ranks
):

    total = len(ranks)


    recall_at_5 = sum(
        rank is not None
        and rank <= TOP_K_FINAL
        for rank in ranks
    ) / total


    hit_at_1 = sum(
        rank == 1
        for rank in ranks
    ) / total


    mrr = sum(
        1 / rank
        if rank is not None
        else 0
        for rank in ranks
    ) / total


    return (
        recall_at_5,
        hit_at_1,
        mrr
    )


# ============================================================
# EVALUATION
# ============================================================

baseline_ranks = []

intent_ranks = []

changed_cases = []


print()

print(
    "=" * 75
)

print(
    "INTENT-AWARE HYBRID EVALUATION"
)

print(
    "=" * 75
)

print()

print(
    "Intent weight:",
    INTENT_WEIGHT
)


for index, (
    question,
    relevant_sources
) in enumerate(
    test_cases,
    start=1
):


    query_embedding = model.encode(
        [question],
        normalize_embeddings=True
    )[0].tolist()


    results = collection.query(
        query_embeddings=[
            query_embedding
        ],
        n_results=TOP_K_RETRIEVAL
    )


    documents = results[
        "documents"
    ][0]

    metadatas = results[
        "metadatas"
    ][0]

    distances = results[
        "distances"
    ][0]


    candidates = []


    for (
        document,
        metadata,
        distance
    ) in zip(
        documents,
        metadatas,
        distances
    ):


        similarity = (
            1 - distance
        )


        lexical = lexical_overlap(
            question,
            document
        )


        hybrid_score = (
            similarity
            +
            LEXICAL_WEIGHT
            * lexical
        )


        bonus = intent_bonus(
            question,
            metadata,
            document
        )


        intent_score = (
            hybrid_score
            +
            bonus
        )


        candidates.append(
            {
                "metadata": metadata,
                "similarity": similarity,
                "lexical": lexical,
                "hybrid": hybrid_score,
                "intent_bonus": bonus,
                "intent_score": intent_score
            }
        )


    # ========================================================
    # BASELINE RANKING
    # ========================================================

    baseline_sorted = sorted(
        candidates,
        key=lambda x: x[
            "hybrid"
        ],
        reverse=True
    )


    baseline_rank = None


    for rank, candidate in enumerate(
        baseline_sorted[
            :TOP_K_FINAL
        ],
        start=1
    ):


        if (
            candidate[
                "metadata"
            ][
                "source_id"
            ]
            in relevant_sources
        ):

            baseline_rank = rank

            break


    # ========================================================
    # INTENT-AWARE RANKING
    # ========================================================

    intent_sorted = sorted(
        candidates,
        key=lambda x: x[
            "intent_score"
        ],
        reverse=True
    )


    intent_rank = None


    for rank, candidate in enumerate(
        intent_sorted[
            :TOP_K_FINAL
        ],
        start=1
    ):


        if (
            candidate[
                "metadata"
            ][
                "source_id"
            ]
            in relevant_sources
        ):

            intent_rank = rank

            break


    baseline_ranks.append(
        baseline_rank
    )

    intent_ranks.append(
        intent_rank
    )


    # ========================================================
    # RECORD ONLY CHANGED CASES
    # ========================================================

    if (
        baseline_rank
        != intent_rank
    ):


        changed_cases.append(
            {
                "index": index,
                "question": question,
                "expected": "|".join(
                    sorted(
                        relevant_sources
                    )
                ),
                "baseline_rank":
                    baseline_rank,
                "intent_rank":
                    intent_rank,
                "intent":
                    detect_intent(
                        question
                    ),
                "baseline_top5": [
                    x[
                        "metadata"
                    ][
                        "source_id"
                    ]
                    for x
                    in baseline_sorted[
                        :5
                    ]
                ],
                "intent_top5": [
                    x[
                        "metadata"
                    ][
                        "source_id"
                    ]
                    for x
                    in intent_sorted[
                        :5
                    ]
                ]
            }
        )


# ============================================================
# FINAL METRICS
# ============================================================

baseline_metrics = calculate_metrics(
    baseline_ranks
)

intent_metrics = calculate_metrics(
    intent_ranks
)


print()

print(
    "=" * 75
)

print(
    "FINAL INTENT-AWARE EVALUATION"
)

print(
    "=" * 75
)

print()

print(
    "                         BASELINE       INTENT-AWARE"
)

print(
    "-" * 75
)


print(
    f"Recall@5               "
    f"{baseline_metrics[0]:.4f}"
    f"          "
    f"{intent_metrics[0]:.4f}"
)


print(
    f"Hit@1                  "
    f"{baseline_metrics[1]:.4f}"
    f"          "
    f"{intent_metrics[1]:.4f}"
)


print(
    f"MRR                    "
    f"{baseline_metrics[2]:.4f}"
    f"          "
    f"{intent_metrics[2]:.4f}"
)


print()

print(
    "-" * 75
)


print(
    "Recall@5 change:",
    f"{intent_metrics[0] - baseline_metrics[0]:+.4f}"
)


print(
    "Hit@1 change:",
    f"{intent_metrics[1] - baseline_metrics[1]:+.4f}"
)


print(
    "MRR change:",
    f"{intent_metrics[2] - baseline_metrics[2]:+.4f}"
)


# ============================================================
# CHANGED CASES
# ============================================================

print()

print(
    "=" * 75
)

print(
    "NON-IDENTICAL CASES"
)

print(
    "=" * 75
)

print()


if not changed_cases:

    print(
        "No ranking changes."
    )


else:

    for case in changed_cases:

        print(
            f"[{case['index']}/55]"
        )

        print(
            "Question:",
            case["question"]
        )

        print(
            "Intent:",
            case["intent"]
        )

        print(
            "Expected:",
            case["expected"]
        )

        print(
            "Baseline rank:",
            case["baseline_rank"]
        )

        print(
            "Intent-aware rank:",
            case["intent_rank"]
        )

        print(
            "Baseline Top-5:",
            " | ".join(
                case[
                    "baseline_top5"
                ]
            )
        )

        print(
            "Intent Top-5:",
            " | ".join(
                case[
                    "intent_top5"
                ]
            )
        )

        print()


print(
    "Total changed cases:",
    len(changed_cases)
)


print()

print(
    "Evaluation completed."
)