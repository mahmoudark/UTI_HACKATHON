import re
import chromadb
from sentence_transformers import SentenceTransformer
from openpyxl import load_workbook


# ============================================================
# SETTINGS
# ============================================================

EXCEL_FILE = "UTI_evaluation_dataset_v2.xlsx"

TOP_K_RETRIEVAL = 10
TOP_K_FINAL = 5

# Conservative lexical bonus.
# The goal is to test whether lexical evidence can improve
# ranking without overpowering semantic similarity.
LEXICAL_WEIGHT = 0.10


# ============================================================
# LOAD DATASET
# ============================================================

print("Loading evaluation dataset...")

wb = load_workbook(
    EXCEL_FILE,
    read_only=True,
    data_only=True
)

ws = wb["Evaluation_Set"]

test_cases = []

for row in ws.iter_rows(min_row=2, values_only=True):

    question = row[1]
    relevant_sources = row[2]

    if question and relevant_sources:

        relevant_sources = set(
            str(relevant_sources).split("|")
        )

        test_cases.append(
            (question, relevant_sources)
        )

print("Questions loaded:", len(test_cases))


# ============================================================
# LOAD MPNet + MPNet DATABASE
# ============================================================

print()
print("Loading MPNet...")

model = SentenceTransformer(
    "all-mpnet-base-v2"
)

print("MPNet loaded")


client = chromadb.PersistentClient(
    path="./chroma_db_mpnet"
)

collection = client.get_collection(
    name="uti_guideline_mpnet"
)

print("MPNet database loaded")
print("Documents:", collection.count())


# ============================================================
# TEXT NORMALIZATION
# ============================================================

STOPWORDS = {
    "the", "a", "an", "and", "or", "of", "to", "for", "with",
    "what", "which", "how", "should", "be", "is", "are", "was",
    "were", "do", "does", "did", "when", "where", "who", "that",
    "this", "these", "those", "in", "on", "at", "from", "by",
    "about", "all", "people", "person", "patients", "patient"
}


def tokenize(text):
    words = re.findall(r"[a-zA-Z0-9]+", text.lower())

    return {
        word
        for word in words
        if len(word) >= 3 and word not in STOPWORDS
    }


def lexical_overlap(query, document):
    """
    F1-style token overlap between the query and document.
    This is deliberately conservative: it is only a small
    ranking signal added to MPNet similarity.
    """

    query_tokens = tokenize(query)
    document_tokens = tokenize(document)

    if not query_tokens or not document_tokens:
        return 0.0

    overlap = len(
        query_tokens.intersection(document_tokens)
    )

    precision = overlap / len(query_tokens)
    recall = overlap / len(document_tokens)

    if precision + recall == 0:
        return 0.0

    return (
        2 * precision * recall
        / (precision + recall)
    )


# ============================================================
# METRICS
# ============================================================

def calculate_metrics(ranks):

    total = len(ranks)

    hits = [
        rank for rank in ranks
        if rank is not None
    ]

    recall_at_5 = (
        len(hits) / total
    )

    hit_at_1 = (
        sum(
            1 for rank in ranks
            if rank == 1
        ) / total
    )

    mrr = (
        sum(
            1 / rank
            for rank in hits
        ) / total
    )

    return recall_at_5, hit_at_1, mrr


# ============================================================
# EVALUATION
# ============================================================

baseline_ranks = []
hybrid_ranks = []

print()
print("=" * 70)
print("MPNET BASELINE VS MPNet + CONSERVATIVE HYBRID")
print("=" * 70)

for index, (question, relevant_sources) in enumerate(
    test_cases,
    start=1
):

    print()
    print(f"[{index}/{len(test_cases)}] {question}")

    query_embedding = model.encode(
        [question],
        normalize_embeddings=True
    )[0].tolist()

    results = collection.query(
        query_embeddings=[query_embedding],
        n_results=TOP_K_RETRIEVAL
    )

    documents = results["documents"][0]
    metadatas = results["metadatas"][0]
    distances = results["distances"][0]

    candidates = []

    for document, metadata, distance in zip(
        documents,
        metadatas,
        distances
    ):

        similarity = 1 - distance

        lexical_score = lexical_overlap(
            question,
            document
        )

        hybrid_score = (
            similarity
            +
            LEXICAL_WEIGHT * lexical_score
        )

        candidates.append({
            "metadata": metadata,
            "similarity": similarity,
            "lexical_score": lexical_score,
            "hybrid_score": hybrid_score
        })

    # --------------------------------------------------------
    # BASELINE: MPNet ranking
    # --------------------------------------------------------

    baseline_rank = None

    for rank, metadata in enumerate(
        metadatas[:TOP_K_FINAL],
        start=1
    ):

        if metadata["source_id"] in relevant_sources:

            baseline_rank = rank
            break

    baseline_ranks.append(
        baseline_rank
    )

    # --------------------------------------------------------
    # HYBRID: MPNet similarity + conservative lexical signal
    # --------------------------------------------------------

    candidates.sort(
        key=lambda item: item["hybrid_score"],
        reverse=True
    )

    hybrid_rank = None

    for rank, candidate in enumerate(
        candidates[:TOP_K_FINAL],
        start=1
    ):

        if candidate["metadata"]["source_id"] in relevant_sources:

            hybrid_rank = rank
            break

    hybrid_ranks.append(
        hybrid_rank
    )

    print(
        f"Baseline rank: {baseline_rank} | "
        f"Hybrid rank: {hybrid_rank}"
    )


# ============================================================
# FINAL RESULTS
# ============================================================

baseline_metrics = calculate_metrics(
    baseline_ranks
)

hybrid_metrics = calculate_metrics(
    hybrid_ranks
)

print()
print()
print("=" * 70)
print("FINAL 55-QUESTION HYBRID EVALUATION")
print("=" * 70)

print()
print("                         BASELINE       HYBRID")
print("-" * 70)

print(
    f"Recall@5               {baseline_metrics[0]:.4f}"
    f"          {hybrid_metrics[0]:.4f}"
)

print(
    f"Hit@1                  {baseline_metrics[1]:.4f}"
    f"          {hybrid_metrics[1]:.4f}"
)

print(
    f"MRR                    {baseline_metrics[2]:.4f}"
    f"          {hybrid_metrics[2]:.4f}"
)

print()
print("-" * 70)

print(
    "Recall@5 change:",
    f"{hybrid_metrics[0] - baseline_metrics[0]:+.4f}"
)

print(
    "Hit@1 change:",
    f"{hybrid_metrics[1] - baseline_metrics[1]:+.4f}"
)

print(
    "MRR change:",
    f"{hybrid_metrics[2] - baseline_metrics[2]:+.4f}"
)

print()
print("Lexical weight:", LEXICAL_WEIGHT)
print("Evaluation completed.")
