import re
import chromadb
from sentence_transformers import SentenceTransformer
from openpyxl import load_workbook


# ============================================================
# SETTINGS
# ============================================================

EXCEL_FILE = "UTI_evaluation_dataset_v2.xlsx"

WEIGHTS = [0.00, 0.05, 0.10, 0.15, 0.20, 0.25]
TOP_K_RETRIEVAL = 10
TOP_K_FINAL = 5


# ============================================================
# LOAD DATASET
# ============================================================

print("Loading evaluation dataset...")

wb = load_workbook(
    EXCEL_FILE,
    read_only=True,
    data_only=True
)

ws = wb["Evaluation_Set"]

test_cases = []

for row in ws.iter_rows(min_row=2, values_only=True):
    question = row[1]
    relevant_sources = row[2]

    if question and relevant_sources:
        test_cases.append(
            (
                question,
                set(str(relevant_sources).split("|"))
            )
        )

print("Questions loaded:", len(test_cases))


# ============================================================
# LOAD MPNet
# ============================================================

print()
print("Loading MPNet...")

model = SentenceTransformer(
    "all-mpnet-base-v2"
)

print("MPNet loaded")


# ============================================================
# LOAD DATABASE
# ============================================================

client = chromadb.PersistentClient(
    path="./chroma_db_mpnet"
)

collection = client.get_collection(
    name="uti_guideline_mpnet"
)

print("Documents:", collection.count())


# ============================================================
# TOKENIZATION
# ============================================================

STOPWORDS = {
    "the", "a", "an", "and", "or", "of", "to", "for", "with",
    "what", "which", "how", "should", "be", "is", "are", "was",
    "were", "do", "does", "did", "when", "where", "who", "that",
    "this", "these", "those", "in", "on", "at", "from", "by",
    "about", "all", "people", "person", "patients", "patient"
}


def tokenize(text):
    words = re.findall(r"[a-zA-Z0-9]+", text.lower())

    return {
        word
        for word in words
        if len(word) >= 3 and word not in STOPWORDS
    }


def lexical_overlap(query, document):
    q = tokenize(query)
    d = tokenize(document)

    if not q or not d:
        return 0.0

    overlap = len(q.intersection(d))

    precision = overlap / len(q)
    recall = overlap / len(d)

    if precision + recall == 0:
        return 0.0

    return 2 * precision * recall / (precision + recall)


# ============================================================
# METRICS
# ============================================================

def metrics(ranks):
    total = len(ranks)

    hits = [
        r for r in ranks
        if r is not None
    ]

    recall5 = len(hits) / total

    hit1 = sum(
        1 for r in ranks
        if r == 1
    ) / total

    mrr = sum(
        1 / r
        for r in hits
    ) / total

    return recall5, hit1, mrr


# ============================================================
# PRECOMPUTE RETRIEVAL
# ============================================================

print()
print("Precomputing MPNet retrieval...")

prepared = []

for index, (question, relevant_sources) in enumerate(
    test_cases,
    start=1
):

    query_embedding = model.encode(
        [question],
        normalize_embeddings=True
    )[0].tolist()

    results = collection.query(
        query_embeddings=[query_embedding],
        n_results=TOP_K_RETRIEVAL
    )

    documents = results["documents"][0]
    metadatas = results["metadatas"][0]
    distances = results["distances"][0]

    candidates = []

    for document, metadata, distance in zip(
        documents,
        metadatas,
        distances
    ):
        similarity = 1 - distance

        lexical_score = lexical_overlap(
            question,
            document
        )

        candidates.append({
            "metadata": metadata,
            "similarity": similarity,
            "lexical_score": lexical_score
        })

    prepared.append(
        (question, relevant_sources, candidates)
    )

    print(
        f"\rPrepared {index}/{len(test_cases)}",
        end=""
    )

print()


# ============================================================
# TUNE WEIGHTS
# ============================================================

results_table = []

print()
print("=" * 70)
print("LEXICAL WEIGHT TUNING")
print("=" * 70)

for weight in WEIGHTS:

    ranks = []

    for question, relevant_sources, candidates in prepared:

        ranked = sorted(
            candidates,
            key=lambda item:
                item["similarity"]
                +
                weight * item["lexical_score"],
            reverse=True
        )

        rank_found = None

        for rank, candidate in enumerate(
            ranked[:TOP_K_FINAL],
            start=1
        ):

            if candidate["metadata"]["source_id"] in relevant_sources:
                rank_found = rank
                break

        ranks.append(rank_found)

    recall5, hit1, mrr = metrics(ranks)

    results_table.append(
        (weight, recall5, hit1, mrr)
    )

    print(
        f"Weight {weight:.2f} | "
        f"Recall@5: {recall5:.4f} | "
        f"Hit@1: {hit1:.4f} | "
        f"MRR: {mrr:.4f}"
    )


# ============================================================
# SELECT BEST WEIGHT
# ============================================================

# Primary objective: maximize MRR.
# Tie-breaker: Hit@1, then Recall@5.
best = max(
    results_table,
    key=lambda row: (
        row[3],
        row[2],
        row[1]
    )
)

print()
print("=" * 70)
print("BEST WEIGHT")
print("=" * 70)

print(
    f"Weight: {best[0]:.2f}"
)

print(
    f"Recall@5: {best[1]:.4f}"
)

print(
    f"Hit@1: {best[2]:.4f}"
)

print(
    f"MRR: {best[3]:.4f}"
)

print()
print("Tuning completed.")
