import statistics
from openpyxl import load_workbook
from answer_engine_v3 import answer_with_guard

EXCEL_FILE = "UTI_evaluation_dataset_v2.xlsx"

wb = load_workbook(EXCEL_FILE, read_only=True, data_only=True)
ws = wb["Evaluation_Set"]

test_cases = []

for row in ws.iter_rows(min_row=2, values_only=True):
    question = row[1]
    relevant_sources = row[2]

    if question and relevant_sources:
        relevant = {x.strip() for x in str(relevant_sources).split("|")}
        test_cases.append((question, relevant))

print("=" * 75)
print("DAY 4 METRICS EVALUATION")
print("=" * 75)
print("Questions:", len(test_cases))

citation_scores = []
grounding_scores = []

answered = 0
refused = 0

for i, (question, relevant) in enumerate(test_cases, start=1):

    result = answer_with_guard(question)

    if result["status"] == "REFUSED":
        refused += 1
        print(f"[{i:02}/{len(test_cases)}] REFUSED | {question}")
        continue

    answered += 1

    source = result.get("source") or {}
    source_id = source.get("source_id")

    citation_correct = source_id in relevant
    citation_scores.append(1.0 if citation_correct else 0.0)

    grounding = result.get("grounding")
    if grounding is not None:
        grounding_scores.append(float(grounding))

    print(
        f"[{i:02}/{len(test_cases)}] "
        f"source={source_id} "
        f"expected={','.join(sorted(relevant))} "
        f"citation={'PASS' if citation_correct else 'FAIL'} "
        f"grounding={grounding}"
    )

print()
print("=" * 75)
print("RESULTS")
print("=" * 75)

citation_accuracy = (
    statistics.mean(citation_scores)
    if citation_scores else 0.0
)

grounding_mean = (
    statistics.mean(grounding_scores)
    if grounding_scores else 0.0
)

print(f"Answered:           {answered}")
print(f"Refused:            {refused}")
print(f"Citation Accuracy:  {citation_accuracy:.4f}")
print(f"Grounding Coverage: {grounding_mean:.4f}")

print()
print("IMPORTANT:")
print(
    "Grounding Coverage is the system's lexical grounding check, "
    "not a claim-level Faithfulness score."
)