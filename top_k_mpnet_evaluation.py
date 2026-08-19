import re
from openpyxl import load_workbook
import chromadb
from sentence_transformers import SentenceTransformer


EXCEL_FILE = "UTI_evaluation_dataset_v2.xlsx"

TOP_K_RETRIEVAL = 10
K_VALUES = [1, 2, 3, 4, 5]
LEXICAL_WEIGHT = 0.10


# ============================================================
# LOAD DATASET
# ============================================================

wb = load_workbook(
    EXCEL_FILE,
    read_only=True,
    data_only=True,
)

ws = wb["Evaluation_Set"]

test_cases = []

for row in ws.iter_rows(min_row=2, values_only=True):
    question = row[1]
    sources = row[2]

    if question and sources:
        relevant = {
            x.strip()
            for x in str(sources).split("|")
        }

        test_cases.append(
            (question, relevant)
        )

print("Questions loaded:", len(test_cases))


# ============================================================
# LOAD MODEL + DATABASE
# ============================================================

print("Loading MPNet...")

model = SentenceTransformer(
    "all-mpnet-base-v2"
)

client = chromadb.PersistentClient(
    path="./chroma_db_mpnet"
)

collection = client.get_collection(
    name="uti_guideline_mpnet"
)

print("Documents:", collection.count())


# ============================================================
# TOKENIZATION
# ============================================================

STOPWORDS = {
    "the", "a", "an", "and", "or", "of", "to",
    "for", "with", "what", "which", "how",
    "should", "be", "is", "are", "was", "were",
    "do", "does", "did", "when", "where",
    "who", "that", "this", "these", "those",
    "in", "on", "at", "from", "by", "about",
    "all", "people", "person", "patients", "patient"
}


def tokens(text):

    return {
        w
        for w in re.findall(
            r"[a-zA-Z0-9]+",
            text.lower()
        )
        if len(w) >= 3
        and w not in STOPWORDS
    }


def lexical_overlap(question, document):

    q = tokens(question)
    d = tokens(document)

    if not q or not d:
        return 0.0

    overlap = len(q & d)

    precision = overlap / len(q)
    recall = overlap / len(d)

    if precision + recall == 0:
        return 0.0

    return (
        2 * precision * recall
        / (precision + recall)
    )


# ============================================================
# PRECOMPUTE RETRIEVAL
# ============================================================

all_results = []

for index, (question, relevant) in enumerate(
    test_cases,
    start=1
):

    embedding = model.encode(
        [question],
        normalize_embeddings=True
    )[0].tolist()

    result = collection.query(
        query_embeddings=[embedding],
        n_results=TOP_K_RETRIEVAL
    )

    candidates = []

    for document, metadata, distance in zip(
        result["documents"][0],
        result["metadatas"][0],
        result["distances"][0]
    ):

        similarity = 1 - distance

        lexical = lexical_overlap(
            question,
            document
        )

        hybrid = (
            similarity
            + LEXICAL_WEIGHT * lexical
        )

        candidates.append({
            "source_id": metadata["source_id"],
            "similarity": similarity,
            "hybrid": hybrid,
        })

    candidates.sort(
        key=lambda x: x["hybrid"],
        reverse=True
    )

    all_results.append(
        (question, relevant, candidates)
    )

    print(
        f"Prepared {index}/{len(test_cases)}"
    )


# ============================================================
# METRICS
# ============================================================

def evaluate_k(k):

    hit_count = 0
    reciprocal_sum = 0.0
    precision_sum = 0.0

    for question, relevant, candidates in all_results:

        top_k = candidates[:k]

        relevant_hits = sum(
            item["source_id"] in relevant
            for item in top_k
        )

        precision_sum += (
            relevant_hits / k
        )

        rank = None

        for position, item in enumerate(
            top_k,
            start=1
        ):

            if item["source_id"] in relevant:
                rank = position
                break

        if rank is not None:

            hit_count += 1
            reciprocal_sum += 1 / rank

    n = len(all_results)

    return {
        "precision": precision_sum / n,
        "recall": hit_count / n,
        "mrr": reciprocal_sum / n,
    }


# ============================================================
# RESULTS
# ============================================================

print()
print("=" * 75)
print("MPNET + HYBRID TOP-K EVALUATION")
print("=" * 75)

print()
print(
    f"{'K':<5}"
    f"{'Precision@K':<18}"
    f"{'Recall@K':<15}"
    f"{'MRR':<15}"
)

print("-" * 75)

results = {}

for k in K_VALUES:

    metrics = evaluate_k(k)

    results[k] = metrics

    print(
        f"{k:<5}"
        f"{metrics['precision']:<18.4f}"
        f"{metrics['recall']:<15.4f}"
        f"{metrics['mrr']:<15.4f}"
    )


# ============================================================
# BEST BY DIFFERENT OBJECTIVES
# ============================================================

best_precision_k = max(
    K_VALUES,
    key=lambda k: results[k]["precision"]
)

best_recall_k = max(
    K_VALUES,
    key=lambda k: results[k]["recall"]
)

best_mrr_k = max(
    K_VALUES,
    key=lambda k: results[k]["mrr"]
)


print()
print("=" * 75)
print("BEST K BY METRIC")
print("=" * 75)

print(
    f"Best Precision@K: K={best_precision_k} "
    f"({results[best_precision_k]['precision']:.4f})"
)

print(
    f"Best Recall@K:    K={best_recall_k} "
    f"({results[best_recall_k]['recall']:.4f})"
)

print(
    f"Best MRR:         K={best_mrr_k} "
    f"({results[best_mrr_k]['mrr']:.4f})"
)

print()
print("Evaluation completed.")