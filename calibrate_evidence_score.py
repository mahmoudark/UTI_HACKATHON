import numpy as np
import chromadb
from sentence_transformers import SentenceTransformer
from openpyxl import load_workbook


EXCEL_FILE = "UTI_evaluation_dataset_v2.xlsx"
MODEL_NAME = "all-mpnet-base-v2"
DATABASE_PATH = "./chroma_db_mpnet"
COLLECTION_NAME = "uti_guideline_mpnet"

TOP_K = 4


# ============================================================
# LOAD DATASET
# ============================================================

wb = load_workbook(
    EXCEL_FILE,
    read_only=True,
    data_only=True
)

ws = wb["Evaluation_Set"]

test_cases = []

for row in ws.iter_rows(min_row=2, values_only=True):

    question = row[1]
    relevant_sources = row[2]

    if question and relevant_sources:
        relevant = {
            x.strip()
            for x in str(relevant_sources).split("|")
        }

        test_cases.append(
            (question, relevant)
        )

print("Questions:", len(test_cases))


# ============================================================
# LOAD MODEL / DATABASE
# ============================================================

print("Loading MPNet...")

model = SentenceTransformer(MODEL_NAME)

client = chromadb.PersistentClient(
    path=DATABASE_PATH
)

collection = client.get_collection(
    name=COLLECTION_NAME
)

print("Documents:", collection.count())


# ============================================================
# BUILD TOP-K LABELED EXAMPLES
# ============================================================

scores = []
labels = []

for i, (question, relevant_sources) in enumerate(
    test_cases,
    start=1
):

    embedding = model.encode(
        [question],
        normalize_embeddings=True
    )[0].tolist()

    result = collection.query(
        query_embeddings=[embedding],
        n_results=TOP_K
    )

    for metadata, distance in zip(
        result["metadatas"][0],
        result["distances"][0]
    ):

        # True cosine for this Chroma collection.
        cosine = 1 - (distance / 2)

        source_id = metadata["source_id"]

        label = (
            1
            if source_id in relevant_sources
            else 0
        )

        scores.append(cosine)
        labels.append(label)

    print(
        f"Prepared {i}/{len(test_cases)}"
    )


scores = np.array(scores, dtype=float)
labels = np.array(labels, dtype=float)


# ============================================================
# BIN-BASED EMPIRICAL CALIBRATION
# ============================================================

print()
print("=" * 80)
print("EVIDENCE SCORE CALIBRATION")
print("=" * 80)

print()
print(
    "This calibration estimates the empirical relevance rate "
    "of retrieved Top-K chunks."
)

bins = np.arange(
    0.50,
    0.951,
    0.05
)

calibration = []

print()
print(
    "Cosine Range       Samples     Relevant     Match Rate"
)

print("-" * 65)

for low in bins:

    high = low + 0.05

    if high > 0.951:
        break

    mask = (
        (scores >= low)
        & (scores < high)
    )

    count = int(mask.sum())

    if count == 0:
        continue

    relevant = int(labels[mask].sum())

    rate = relevant / count

    calibration.append(
        {
            "low": float(low),
            "high": float(high),
            "samples": count,
            "relevant": relevant,
            "rate": rate
        }
    )

    print(
        f"{low:.2f} - {high:.2f}       "
        f"{count:<11d}"
        f"{relevant:<11d}"
        f"{rate * 100:6.2f}%"
    )


# ============================================================
# FIND 90% EVIDENCE THRESHOLD
# ============================================================

print()
print("=" * 80)
print("90% EVIDENCE MATCH CANDIDATES")
print("=" * 80)

candidates = [
    x
    for x in calibration
    if x["samples"] >= 5
    and x["rate"] >= 0.90
]

if candidates:

    for item in candidates:

        midpoint = (
            item["low"] + item["high"]
        ) / 2

        print(
            f"Cosine >= ~{midpoint:.2f} | "
            f"Match rate = {item['rate'] * 100:.2f}% | "
            f"Samples = {item['samples']}"
        )

else:

    print(
        "No cosine range reached 90% empirical relevance "
        "with at least 5 samples."
    )


# ============================================================
# LOOK AT CURRENT PRODUCTION EXAMPLE
# ============================================================

example_query = (
    "For someone with lower UTI, "
    "what is recommended for pain relief?"
)

embedding = model.encode(
    [example_query],
    normalize_embeddings=True
)[0].tolist()

result = collection.query(
    query_embeddings=[embedding],
    n_results=TOP_K
)

print()
print("=" * 80)
print("CURRENT EXAMPLE")
print("=" * 80)

for rank, (metadata, distance) in enumerate(
    zip(
        result["metadatas"][0],
        result["distances"][0]
    ),
    start=1
):

    cosine = 1 - (distance / 2)

    matching_bucket = None

    for item in calibration:

        if (
            cosine >= item["low"]
            and cosine < item["high"]
        ):
            matching_bucket = item
            break

    if matching_bucket:

        evidence_score = (
            matching_bucket["rate"]
        )

    else:

        evidence_score = None

    print(
        f"Rank {rank} | "
        f"Source={metadata['source_id']} | "
        f"Cosine={cosine:.4f} | "
        f"Evidence Match="
        f"{'N/A' if evidence_score is None else f'{evidence_score * 100:.2f}%'}"
    )


print()
print("Calibration completed.")