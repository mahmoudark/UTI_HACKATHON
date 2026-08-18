import re
import chromadb
from sentence_transformers import SentenceTransformer
from openpyxl import load_workbook
from intent_classifier import detect_intent

# ============================================================
# SETTINGS
# ============================================================

EXCEL_FILE = "UTI_evaluation_dataset_v2.xlsx"

MODEL_NAME = "all-mpnet-base-v2"

DATABASE_PATH = "./chroma_db_mpnet"

COLLECTION_NAME = "uti_guideline_mpnet"

TOP_K_RETRIEVAL = 10

TOP_K_FINAL = 5

LEXICAL_WEIGHT = 0.10


# ============================================================
# LOAD DATASET
# ============================================================

print("Loading evaluation dataset...")

wb = load_workbook(
    EXCEL_FILE,
    read_only=True,
    data_only=True
)

ws = wb["Evaluation_Set"]

test_cases = []

for row in ws.iter_rows(
    min_row=2,
    values_only=True
):

    question = row[1]
    expected = row[2]

    if question and expected:

        relevant_sources = set(
            str(expected).split("|")
        )

        test_cases.append(
            (
                question,
                relevant_sources
            )
        )

print(
    "Questions loaded:",
    len(test_cases)
)


# ============================================================
# LOAD MODEL
# ============================================================

print()
print("Loading MPNet...")

model = SentenceTransformer(
    MODEL_NAME
)

print("MPNet loaded")


client = chromadb.PersistentClient(
    path=DATABASE_PATH
)

collection = client.get_collection(
    name=COLLECTION_NAME
)

print(
    "Documents:",
    collection.count()
)


# ============================================================
# TOKENIZATION
# ============================================================

STOPWORDS = {
    "the", "a", "an", "and", "or", "of", "to", "for", "with",
    "what", "which", "how", "should", "be", "is", "are", "was",
    "were", "do", "does", "did", "when", "where", "who", "that",
    "this", "these", "those", "in", "on", "at", "from", "by",
    "about", "all", "people", "person", "patients", "patient"
}


def tokenize(text):

    return {
        word
        for word in re.findall(
            r"[a-zA-Z0-9]+",
            text.lower()
        )
        if len(word) >= 3
        and word not in STOPWORDS
    }


def lexical_overlap(
    query,
    document
):

    query_tokens = tokenize(query)

    document_tokens = tokenize(document)

    if (
        not query_tokens
        or not document_tokens
    ):
        return 0.0

    overlap = len(
        query_tokens.intersection(
            document_tokens
        )
    )

    precision = (
        overlap
        / len(query_tokens)
    )

    recall = (
        overlap
        / len(document_tokens)
    )

    if precision + recall == 0:
        return 0.0

    return (
        2 * precision * recall
        / (precision + recall)
    )


# ============================================================
# INTENT DETECTION
# ============================================================




# ============================================================
# INTENT BONUS
# ============================================================

def intent_bonus(
    query,
    metadata,
    document,
    intent_weight
):

    intent = detect_intent(
        query
    )

    source_id = str(
        metadata.get(
            "source_id",
            ""
        )
    )

    document_lower = document.lower()

    if intent == "ANTIBIOTIC_SELECTION":

        if source_id == "1.4.1":
            return intent_weight

        if (
            "when prescribing antibiotic treatment"
            in document_lower
        ):
            return intent_weight * 0.50

        if (
            "local antimicrobial resistance data"
            in document_lower
        ):
            return intent_weight * 0.50

    if intent == "TABLE_1":

        if source_id == "TABLE_1":
            return intent_weight

    if intent == "TABLE_2":

        if source_id == "TABLE_2":
            return intent_weight

    if intent == "TABLE_3":

        if source_id == "TABLE_3":
            return intent_weight

    if intent == "TABLE_4":

        if source_id == "TABLE_4":
            return intent_weight

    return 0.0


# ============================================================
# PRECOMPUTE ALL QUESTIONS
# ============================================================

print()
print("Precomputing retrieval results...")

prepared = []

for index, (
    question,
    relevant_sources
) in enumerate(
    test_cases,
    start=1
):

    embedding = model.encode(
        [question],
        normalize_embeddings=True
    )[0].tolist()

    results = collection.query(
        query_embeddings=[embedding],
        n_results=TOP_K_RETRIEVAL
    )

    documents = results[
        "documents"
    ][0]

    metadatas = results[
        "metadatas"
    ][0]

    distances = results[
        "distances"
    ][0]

    candidates = []

    for (
        document,
        metadata,
        distance
    ) in zip(
        documents,
        metadatas,
        distances
    ):

        similarity = 1 - distance

        lexical = lexical_overlap(
            question,
            document
        )

        hybrid = (
            similarity
            +
            LEXICAL_WEIGHT
            * lexical
        )

        candidates.append({
            "metadata": metadata,
            "document": document,
            "similarity": similarity,
            "lexical": lexical,
            "hybrid": hybrid
        })

    prepared.append({
        "question": question,
        "expected": relevant_sources,
        "candidates": candidates
    })

    print(
        f"Prepared {index}/{len(test_cases)}"
    )


# ============================================================
# ROBUSTNESS QUESTIONS
# ============================================================

ROBUSTNESS_CASES = [

    "What should be considered when prescribing antibiotics for lower UTI?",

    "What factors should be considered when choosing antibiotics for lower UTI?",

    "What should clinicians consider when selecting an antibiotic for lower UTI?",

    "What factors should guide antibiotic prescribing for lower UTI?",

    "What should be taken into account when prescribing antibiotics for lower UTI?",

    "What considerations are important when prescribing antibiotics for lower UTI?",

    "What factors should be considered before prescribing antibiotic treatment for lower UTI?",

    "What should be considered when choosing an antibiotic for a lower UTI?",

    "What factors influence the choice of antibiotic for lower UTI?",

    "How should antibiotic prescribing decisions be made for lower UTI?"
]


# ============================================================
# RETRIEVE SINGLE QUESTION
# ============================================================

def retrieve_question(
    question,
    intent_weight
):

    embedding = model.encode(
        [question],
        normalize_embeddings=True
    )[0].tolist()

    results = collection.query(
        query_embeddings=[embedding],
        n_results=TOP_K_RETRIEVAL
    )

    documents = results[
        "documents"
    ][0]

    metadatas = results[
        "metadatas"
    ][0]

    distances = results[
        "distances"
    ][0]

    candidates = []

    for (
        document,
        metadata,
        distance
    ) in zip(
        documents,
        metadatas,
        distances
    ):

        similarity = 1 - distance

        lexical = lexical_overlap(
            question,
            document
        )

        hybrid = (
            similarity
            +
            LEXICAL_WEIGHT
            * lexical
        )

        bonus = intent_bonus(
            question,
            metadata,
            document,
            intent_weight
        )

        candidates.append({
            "source_id":
                metadata["source_id"],
            "score":
                hybrid + bonus
        })

    candidates.sort(
        key=lambda x: x["score"],
        reverse=True
    )

    return candidates


# ============================================================
# EVALUATE BENCHMARK
# ============================================================

def evaluate_benchmark(
    intent_weight
):

    ranks = []

    for item in prepared:

        question = item[
            "question"
        ]

        expected = item[
            "expected"
        ]

        candidates = []

        for candidate in item[
            "candidates"
        ]:

            bonus = intent_bonus(
                question,
                candidate["metadata"],
                candidate["document"],
                intent_weight
            )

            candidates.append({
                "source_id":
                    candidate[
                        "metadata"
                    ][
                        "source_id"
                    ],

                "score":
                    candidate[
                        "hybrid"
                    ]
                    + bonus
            })

        candidates.sort(
            key=lambda x: x["score"],
            reverse=True
        )

        rank = None

        for position, candidate in enumerate(
            candidates[:TOP_K_FINAL],
            start=1
        ):

            if candidate[
                "source_id"
            ] in expected:

                rank = position

                break

        ranks.append(
            rank
        )

    total = len(ranks)

    recall = sum(
        rank is not None
        for rank in ranks
    ) / total

    hit1 = sum(
        rank == 1
        for rank in ranks
    ) / total

    mrr = sum(
        1 / rank
        if rank is not None
        else 0
        for rank in ranks
    ) / total

    return recall, hit1, mrr


# ============================================================
# EVALUATE ROBUSTNESS
# ============================================================

def evaluate_robustness(
    intent_weight
):

    passed = 0

    for question in ROBUSTNESS_CASES:

        candidates = retrieve_question(
            question,
            intent_weight
        )

        rank = None

        for position, candidate in enumerate(
            candidates[:TOP_K_FINAL],
            start=1
        ):

            if candidate[
                "source_id"
            ] == "1.4.1":

                rank = position

                break

        if (
            detect_intent(question)
            == "ANTIBIOTIC_SELECTION"
            and rank == 1
        ):

            passed += 1

    return passed


# ============================================================
# WEIGHT TUNING
# ============================================================

WEIGHTS = [
    0.08,
    0.10,
    0.12,
    0.15,
    0.18,
    0.20
]


print()
print("=" * 75)
print("INTENT WEIGHT + ROBUSTNESS TUNING")
print("=" * 75)

print()

print(
    "Weight       Recall@5     Hit@1       MRR       Robustness"
)

print("-" * 75)


best = None


for weight in WEIGHTS:

    recall, hit1, mrr = evaluate_benchmark(
        weight
    )

    robustness = evaluate_robustness(
        weight
    )

    print(
        f"{weight:<12.2f}"
        f"{recall:<13.4f}"
        f"{hit1:<12.4f}"
        f"{mrr:<11.4f}"
        f"{robustness}/10"
    )

    # Only consider weights that preserve
    # the current benchmark improvement.
    if (
        recall >= 1.0000
        and hit1 >= 0.9818
        and mrr >= 0.9909
    ):

        if best is None:

            best = (
                weight,
                recall,
                hit1,
                mrr,
                robustness
            )

        else:

            if robustness > best[4]:

                best = (
                    weight,
                    recall,
                    hit1,
                    mrr,
                    robustness
                )


# ============================================================
# FINAL RECOMMENDATION
# ============================================================

print()
print("=" * 75)
print("BEST CANDIDATE")
print("=" * 75)

if best is None:

    print()
    print(
        "No tested weight preserved the "
        "current benchmark performance."
    )

else:

    print()
    print(
        "Weight:",
        best[0]
    )

    print(
        "Recall@5:",
        f"{best[1]:.4f}"
    )

    print(
        "Hit@1:",
        f"{best[2]:.4f}"
    )

    print(
        "MRR:",
        f"{best[3]:.4f}"
    )

    print(
        "Robustness:",
        f"{best[4]}/10"
    )

    if best[4] == 10:

        print()
        print(
            "FINAL DECISION: ACCEPT"
        )

    else:

        print()
        print(
            "FINAL DECISION: DO NOT MERGE YET"
        )

print()
print(
    "Tuning completed."
)