import chromadb
from sentence_transformers import SentenceTransformer
from openpyxl import load_workbook


EXCEL_FILE = "UTI_evaluation_dataset_v2.xlsx"

MODEL_NAME = "all-mpnet-base-v2"
DATABASE_PATH = "./chroma_db_mpnet"
COLLECTION_NAME = "uti_guideline_mpnet"

TOP_K_RETRIEVAL = 10
TOP_K_FINAL = 5


# ============================================================
# LOAD DATASET
# ============================================================

wb = load_workbook(
    EXCEL_FILE,
    read_only=True,
    data_only=True
)

ws = wb["Evaluation_Set"]

test_cases = []

for row in ws.iter_rows(min_row=2, values_only=True):
    question = row[1]
    relevant_sources = row[2]

    if question and relevant_sources:
        relevant = {
            x.strip()
            for x in str(relevant_sources).split("|")
        }
        test_cases.append(
            (question, relevant)
        )

print("Questions loaded:", len(test_cases))


# ============================================================
# LOAD MODEL
# ============================================================

print("Loading MPNet...")

model = SentenceTransformer(MODEL_NAME)

print("MPNet loaded")


# ============================================================
# LOAD DATABASE
# ============================================================

client = chromadb.PersistentClient(
    path=DATABASE_PATH
)

collection = client.get_collection(
    name=COLLECTION_NAME
)

print("Documents:", collection.count())


# ============================================================
# PRECOMPUTE COSINE RETRIEVAL
# ============================================================

prepared = []

for i, (question, relevant) in enumerate(
    test_cases,
    start=1
):

    embedding = model.encode(
        [question],
        normalize_embeddings=True
    )[0].tolist()

    results = collection.query(
        query_embeddings=[embedding],
        n_results=TOP_K_RETRIEVAL
    )

    candidates = []

    for document, metadata, distance in zip(
        results["documents"][0],
        results["metadatas"][0],
        results["distances"][0]
    ):

        similarity = 1 - (distance / 2)

        candidates.append({
            "source_id": metadata["source_id"],
            "similarity": similarity
        })

    # COSINE ONLY:
    candidates.sort(
        key=lambda x: x["similarity"],
        reverse=True
    )

    prepared.append(
        (question, relevant, candidates)
    )

    print(
        f"Prepared {i}/{len(test_cases)}"
    )


# ============================================================
# EVALUATION
# ============================================================

def evaluate_threshold(threshold):

    answered = 0
    correct = 0
    ranks = []

    for question, relevant, candidates in prepared:

        selected = [
            item
            for item in candidates
            if item["similarity"] >= threshold
        ]

        selected = selected[:TOP_K_FINAL]

        if not selected:
            ranks.append(None)
            continue

        answered += 1

        found_rank = None

        for rank, item in enumerate(
            selected,
            start=1
        ):

            if item["source_id"] in relevant:
                found_rank = rank
                break

        ranks.append(found_rank)

        if found_rank is not None:
            correct += 1

    total = len(prepared)

    recall = correct / total
    coverage = answered / total

    hit1 = sum(
        1
        for rank in ranks
        if rank == 1
    ) / total

    mrr = sum(
        1 / rank
        for rank in ranks
        if rank is not None
    ) / total

    precision = 0.0

    precision_values = []

    for question, relevant, candidates in prepared:

        selected = [
            item
            for item in candidates
            if item["similarity"] >= threshold
        ][:TOP_K_FINAL]

        if not selected:
            continue

        relevant_count = sum(
            item["source_id"] in relevant
            for item in selected
        )

        precision_values.append(
            relevant_count / len(selected)
        )

    if precision_values:
        precision = sum(precision_values) / len(
            precision_values
        )

    return (
        precision,
        recall,
        coverage,
        hit1,
        mrr
    )


# ============================================================
# RUN
# ============================================================

print()
print("=" * 80)
print("MPNET + COSINE-ONLY THRESHOLD EVALUATION")
print("=" * 80)

print()
print(
    "Threshold      Precision@5   Recall@5     Coverage     Hit@1       MRR"
)

print("-" * 80)

thresholds = [
    0.50,
    0.55,
    0.60,
    0.65,
    0.70,
    0.75,
    0.80,
    0.85,
    0.90
]

safe = []

for threshold in thresholds:

    precision, recall, coverage, hit1, mrr = (
        evaluate_threshold(threshold)
    )

    print(
        f"{threshold:<14.2f}"
        f"{precision:<15.4f}"
        f"{recall:<13.4f}"
        f"{coverage:<13.4f}"
        f"{hit1:<12.4f}"
        f"{mrr:.4f}"
    )

    if recall >= 1.0:
        safe.append(
            (
                threshold,
                precision,
                recall,
                coverage,
                hit1,
                mrr
            )
        )


print()
print("=" * 80)
print("COSINE-ONLY THRESHOLDS WITH 100% RECALL@5")
print("=" * 80)

if safe:

    for item in safe:
        print(
            f"Threshold {item[0]:.2f} | "
            f"Precision {item[1]:.4f} | "
            f"Coverage {item[3]:.4f} | "
            f"Hit@1 {item[4]:.4f} | "
            f"MRR {item[5]:.4f}"
        )

else:

    print(
        "No tested threshold preserved 100% Recall@5."
    )

print()
print("Evaluation completed.")
